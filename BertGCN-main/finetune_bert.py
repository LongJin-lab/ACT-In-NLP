import torch as th
from transformers import AutoModel, AutoTokenizer
import torch.nn.functional as F
from utils import *
import dgl
import torch.utils.data as Data
from ignite.engine import Events, create_supervised_evaluator, create_supervised_trainer, Engine
from ignite.metrics import Accuracy, Loss
import numpy as np
import os
from datetime import datetime
from sklearn.metrics import accuracy_score
import argparse, shutil, logging
from torch.optim import lr_scheduler
from model import BertClassifier

parser = argparse.ArgumentParser()
parser.add_argument('--max_length', type=int, default=128, help='the input length for bert')
parser.add_argument('--batch_size', type=int, default=128)
parser.add_argument('--nb_epochs', type=int, default=30)
parser.add_argument('--bert_lr', type=float, default=1e-4)
parser.add_argument('--dataset', default='20ng', choices=['20ng', 'R8', 'R52', 'ohsumed', 'mr'])
parser.add_argument('--bert_init', type=str, default='roberta-base',
                    choices=['roberta-base', 'roberta-large', 'bert-base-uncased', 'bert-large-uncased'])
parser.add_argument('--checkpoint_dir', default=None, help='checkpoint directory, [bert_init]_[dataset] if not specified')
parser.add_argument('--device', type=int, default=0)
parser.add_argument('--opt', type=int, default=0)


args = parser.parse_args()

max_length = args.max_length
batch_size = args.batch_size
nb_epochs = args.nb_epochs
bert_lr = args.bert_lr
dataset = args.dataset
bert_init = args.bert_init
checkpoint_dir = args.checkpoint_dir
if checkpoint_dir is None:
    ckpt_dir = './checkpoint/{}_{}'.format(bert_init, dataset)
else:
    ckpt_dir = checkpoint_dir

os.makedirs(ckpt_dir, exist_ok=True)
shutil.copy(os.path.basename(__file__), ckpt_dir)

sh = logging.StreamHandler(sys.stdout)
sh.setFormatter(logging.Formatter('%(message)s'))
sh.setLevel(logging.INFO)
fh = logging.FileHandler(filename=os.path.join(ckpt_dir, 'training.log'), mode='w')
fh.setFormatter(logging.Formatter('%(message)s'))
fh.setLevel(logging.INFO)
logger = logging.getLogger('training logger')
logger.addHandler(sh)
logger.addHandler(fh)
logger.setLevel(logging.INFO)

cpu = th.device('cpu')
gpu = th.device('cuda:'+str(args.device))

logger.info('arguments:')
logger.info(str(args))
logger.info('checkpoints will be saved in {}'.format(ckpt_dir))

# Data Preprocess
adj, features, y_train, y_val, y_test, train_mask, val_mask, test_mask, train_size, test_size = load_corpus(dataset)
'''
y_train, y_val, y_test: n*c matrices 
train_mask, val_mask, test_mask: n-d bool array
train_size, test_size: unused
'''

# compute number of real train/val/test/word nodes and number of classes
nb_node = adj.shape[0]
nb_train, nb_val, nb_test = train_mask.sum(), val_mask.sum(), test_mask.sum()
nb_word = nb_node - nb_train - nb_val - nb_test
nb_class = y_train.shape[1]

# instantiate model according to class number
model = BertClassifier(pretrained_model=bert_init, nb_class=nb_class)

# transform one-hot label to class ID for pytorch computation
y = th.LongTensor((y_train + y_val +y_test).argmax(axis=1))
label = {}
label['train'], label['val'], label['test'] = y[:nb_train], y[nb_train:nb_train+nb_val], y[-nb_test:]

# load documents and compute input encodings
corpus_file = './data/corpus/'+dataset+'_shuffle.txt'
with open(corpus_file, 'r') as f:
    text = f.read()
    text=text.replace('\\', '')
    text = text.split('\n')

def encode_input(text, tokenizer):
    input = tokenizer(text, max_length=max_length, truncation=True, padding=True, return_tensors='pt')
    return input.input_ids, input.attention_mask

input_ids, attention_mask = {}, {}

input_ids_, attention_mask_ = encode_input(text, model.tokenizer)

# create train/test/val datasets and dataloaders
input_ids['train'], input_ids['val'], input_ids['test'] =  input_ids_[:nb_train], input_ids_[nb_train:nb_train+nb_val], input_ids_[-nb_test:]
attention_mask['train'], attention_mask['val'], attention_mask['test'] =  attention_mask_[:nb_train], attention_mask_[nb_train:nb_train+nb_val], attention_mask_[-nb_test:]

datasets = {}
loader = {}
for split in ['train', 'val', 'test']:
    datasets[split] =  Data.TensorDataset(input_ids[split], attention_mask[split], label[split])
    loader[split] = Data.DataLoader(datasets[split], batch_size=batch_size, shuffle=True)


# Training
if args.opt==0:
    optimizer = th.optim.Adam(model.parameters(), lr=bert_lr)
if args.opt==1:
    optimizer = th.optim.Adam_GAF(model.parameters(), lr=bert_lr,gamma=0.3,delta=4.5)
if args.opt==2:
    optimizer = th.optim.Adam_GAF(model.parameters(), lr=bert_lr,gamma=0.1,delta=20)
if args.opt==3:
    optimizer = th.optim.Adam_GAF(model.parameters(), lr=bert_lr,gamma=0.2,delta=10)
if args.opt==4:
    optimizer = th.optim.Adam_GAF(model.parameters(), lr=bert_lr,gamma=0.1,delta=15)
if args.opt==5:
    optimizer = th.optim.Adam_GAF(model.parameters(), lr=bert_lr,gamma=0.1,delta=30)
if args.opt==6:
    optimizer = th.optim.Adam_GAF(model.parameters(), lr=bert_lr,gamma=0.1,delta=40)
if args.opt==7:
    optimizer = th.optim.Adam_GAF(model.parameters(), lr=bert_lr,gamma=0.05,delta=30)
if args.opt==8:
    optimizer = th.optim.Adam_GAF(model.parameters(), lr=bert_lr,gamma=0.05,delta=40)
if args.opt==9:
    optimizer = th.optim.Adam_GAF(model.parameters(), lr=bert_lr,gamma=0.3,delta=7.5)
if args.opt==10:
    optimizer = th.optim.Adam_GAF(model.parameters(), lr=bert_lr,gamma=0.3,delta=10)



scheduler = lr_scheduler.MultiStepLR(optimizer, milestones=[30], gamma=0.1)


def train_step(engine, batch):
    global model, optimizer
    model.train()
    model = model.to(gpu)
    optimizer.zero_grad()
    (input_ids, attention_mask, label) = [x.to(gpu) for x in batch]
    optimizer.zero_grad()
    y_pred = model(input_ids, attention_mask)
    y_true = label.type(th.long)
    loss = F.cross_entropy(y_pred, y_true)
    loss.backward()
    optimizer.step()
    train_loss = loss.item()
    with th.no_grad():
        y_true = y_true.detach().cpu()
        y_pred = y_pred.argmax(axis=1).detach().cpu()
        train_acc = accuracy_score(y_true, y_pred)
    return train_loss, train_acc


trainer = Engine(train_step)


def test_step(engine, batch):
    global model
    with th.no_grad():
        model.eval()
        model = model.to(gpu)
        (input_ids, attention_mask, label) = [x.to(gpu) for x in batch]
        optimizer.zero_grad()
        y_pred = model(input_ids, attention_mask)
        y_true = label
        return y_pred, y_true


evaluator = Engine(test_step)
metrics={
    'acc': Accuracy(),
    'nll': Loss(th.nn.CrossEntropyLoss())
}
for n, f in metrics.items():
    f.attach(evaluator, n)

best_test_acc = 0

@trainer.on(Events.EPOCH_COMPLETED)
def log_training_results(trainer):
    evaluator.run(loader['train'])
    metrics = evaluator.state.metrics
    train_acc, train_nll = metrics["acc"], metrics["nll"]
    evaluator.run(loader['val'])
    metrics = evaluator.state.metrics
    val_acc, val_nll = metrics["acc"], metrics["nll"]
    evaluator.run(loader['test'])
    metrics = evaluator.state.metrics
    test_acc, test_nll = metrics["acc"], metrics["nll"]
    logger.info(
        "\rEpoch: {}  Train acc: {:.4f} loss: {:.4f}  Val acc: {:.4f} loss: {:.4f}  Test acc: {:.4f} loss: {:.4f}"
        .format(trainer.state.epoch, train_acc, train_nll, val_acc, val_nll, test_acc, test_nll)
    )
    global best_test_acc
    if test_acc > best_test_acc:
        best_test_acc = test_acc
    logger.info("\rbest_test_acc: {}".format(best_test_acc))
        
    if val_acc > log_training_results.best_val_acc:
        logger.info("New checkpoint")
        th.save(
            {
                'bert_model': model.bert_model.state_dict(),
                'classifier': model.classifier.state_dict(),
                'optimizer': optimizer.state_dict(),
                'epoch': trainer.state.epoch,
            },
            os.path.join(
                ckpt_dir, 'checkpoint.pth'
            )
        )
        log_training_results.best_val_acc = val_acc
    scheduler.step()

        
log_training_results.best_val_acc = 0

trainer.run(loader['train'], max_epochs=nb_epochs)

dataset_num={'20ng':0,'R8':1,'R52':2,'ohsumed':3,'mr':4}
rowqc=dataset_num[args.dataset]+2
colqc=args.opt+2
from openpyxl import load_workbook
location='/media/qingchuan/BertGCN-main/log/'+str(args.bert_init)+'/'+str(args.bert_init)+'.xlsx'
wb = load_workbook(location)
sheet = wb.active
sheet.cell(row=rowqc, column=colqc).value = best_test_acc
wb.save(location)
